import numpy as np
import cv2
import pytesseract
import pywhatkit
from pyautogui import KEYBOARD_KEYS
from openpyxl import Workbook, load_workbook

pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'

image = cv2.imread('C:\\Users\\HP\\OneDrive\\Desktop\\HackOdisha2.0\\image4.jpg')

cv2.imshow("Original Image", image)

gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
#cv2.imshow("1 - Grayscale Conversion", gray)

gray = cv2.bilateralFilter(gray, 11, 17, 17)
#cv2.imshow("2 - Bilateral Filter", gray)

edged = cv2.Canny(gray, 170, 200)
#cv2.imshow("4 - Canny Edges", edged)

(cnts, _) = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
cnts=sorted(cnts, key = cv2.contourArea, reverse = True)[:30] 
NumberPlateCnt = None 

count = 0
for c in cnts:
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.02 * peri, True)
        if len(approx) == 4:  
            NumberPlateCnt = approx 
            break

# Masking the part other than the number plate
mask = np.zeros(gray.shape,np.uint8)
new_image = cv2.drawContours(mask,[NumberPlateCnt],0,255,-1)
new_image = cv2.bitwise_and(image,image,mask=mask)
cv2.namedWindow("Final_image",cv2.WINDOW_NORMAL)
cv2.imshow("Final_image",new_image)

# Configuration for tesseract
config = ('-l eng --oem 1 --psm 7')

# Run tesseract OCR on image
text = pytesseract.image_to_string(new_image, config=config)

# Print recognized text
print(text)


#for tickets
wb = load_workbook('C:\\Users\\HP\\OneDrive\\Desktop\\HackOdisha2.0\\tickets.xlsx')
ws = wb.active
data = text
ws.append([data])
wb.save('C:\\Users\\HP\\OneDrive\\Desktop\\HackOdisha2.0\\tickets.xlsx')

#for only recent number plate
wb1 = load_workbook('C:\\Users\\HP\\OneDrive\\Desktop\\HackOdisha2.0\\oneplate.xlsx')
ws1 = wb1.active
sheet = wb1['Sheet1']
sheet.delete_rows(1)
wb1.save('C:\\Users\\HP\\OneDrive\\Desktop\\HackOdisha2.0\\oneplate.xlsx')
wb1 = load_workbook('C:\\Users\\HP\\OneDrive\\Desktop\\HackOdisha2.0\\oneplate.xlsx')
ws1 = wb1.active
data = text
ws1.append([data])
wb1.save('C:\\Users\\HP\\OneDrive\\Desktop\\HackOdisha2.0\\oneplate.xlsx')


wb3 = load_workbook('C:\\Users\\HP\\OneDrive\\Desktop\\HackOdisha2.0\\vehicle_details.xlsx')
ws3 = wb3.active
wb4 = load_workbook('C:\\Users\\HP\\OneDrive\\Desktop\\HackOdisha2.0\\oneplate.xlsx')
ws4 = wb4.active

for row in range(2,12):
    if (ws4['A1'].value).rstrip() == (ws3['A' + str(row)].value).rstrip(): 
        print('+91'+str(ws3['C' + str(row)].value))
        q ='+91'+str(ws3['C' + str(row)].value)


pywhatkit.sendwhatmsg_instantly(q,'Car in No Parking , INR500 Fine Your acc is debited ')
KEYBOARD_KEYS('enter')

cv2.waitKey(0)

